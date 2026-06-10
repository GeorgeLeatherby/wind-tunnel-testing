"""
Excel writer for the test matrices (openpyxl).

Writes one worksheet with a styled header row followed by the data rows. Task 1
and Task 2 share the same writer; the only difference is whether the wake-probe
X/Y/Z columns are included (Task 2 only).

The column order mirrors the assignment's requested fields:
    Test number | Requested wind speed | Blade pitch | RPM | Yaw |
    Reynolds | TSR | Expected C_P | Expected C_T |
    [Task 2: Wake Probe X / Y / Z] | (derived loads) | (blank measured channels)
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .scaling_report import ScalingFactorRow
from .test_matrix_row import G1_MEASUREMENT_COLUMNS, TestMatrixRow


# Header styling.
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Numeric display formats per column (Excel number-format strings).
_FMT_SPEED = "0.00"
_FMT_PITCH = "0.0"
_FMT_RPM = "0"
_FMT_YAW = "0"
_FMT_RE = "0"
_FMT_TSR = "0.000"
_FMT_COEFF = "0.0000"
_FMT_POWER = "0.00"
_FMT_TORQUE = "0.000"
_FMT_THRUST = "0.00"
_FMT_POS = "0.0"


class TestMatrixExcelWriter:
    """Writes a list of TestMatrixRow into a styled .xlsx file."""

    def __init__(self, include_wake_position: bool) -> None:
        self._include_wake_position = include_wake_position

    # ----------------------------------------------------------------------
    # Column schema: list of (header, attribute-or-key, number-format).
    # For measurement channels the "attribute" is the channel name (looked up in
    # row.measurements).
    # ----------------------------------------------------------------------
    def _columns(self) -> list[tuple[str, str, str | None]]:
        columns: list[tuple[str, str, str | None]] = [
            ("Test number", "test_number", _FMT_RPM),
            ("Requested wind speed [m/s]", "requested_tunnel_speed", _FMT_SPEED),
            ("Equivalent free-air U' [m/s]", "equivalent_free_air_speed", _FMT_SPEED),
            ("Blade pitch [deg]", "blade_pitch_deg", _FMT_PITCH),
            ("Rotor speed [rpm]", "rotor_speed_rpm", _FMT_RPM),
            ("Yaw [deg]", "yaw_deg", _FMT_YAW),
            ("Reynolds [-]", "reynolds", _FMT_RE),
            ("TSR [-]", "tsr", _FMT_TSR),
            ("Expected C_P [-]", "expected_cp", _FMT_COEFF),
            ("Expected C_T (thrust) [-]", "expected_ct_thrust", _FMT_COEFF),
        ]
        if self._include_wake_position:
            columns.extend(
                [
                    ("Wake Probe X [mm]", "probe_x", _FMT_POS),
                    ("Wake Probe Y [mm]", "probe_y", _FMT_POS),
                    ("Wake Probe Z [mm]", "probe_z", _FMT_POS),
                ]
            )
        columns.extend(
            [
                ("Expected Power [W]", "expected_power_w", _FMT_POWER),
                ("Expected Torque [Nm]", "expected_torque_nm", _FMT_TORQUE),
                ("Expected Thrust [N]", "expected_thrust_n", _FMT_THRUST),
            ]
        )
        # Blank measured-channel columns (no number format; values are empty).
        for channel in G1_MEASUREMENT_COLUMNS:
            columns.append((channel, channel, None))
        return columns

    # ----------------------------------------------------------------------
    def _cell_value(self, row: TestMatrixRow, attribute: str):
        """Resolve a column value from a row: real attribute or measurement key."""
        if hasattr(row, attribute):
            return getattr(row, attribute)
        # Otherwise it is a blank measurement channel.
        return row.measurements[attribute]

    # ----------------------------------------------------------------------
    def write(
        self,
        rows: list[TestMatrixRow],
        sheet_title: str,
        file_path: str,
        scaling_rows: list[ScalingFactorRow] | None = None,
    ) -> None:
        """Create the workbook, write header + rows, style it and save.

        If `scaling_rows` is given, a second "Scaling (model/full-scale)" sheet is
        added that lists the applied Lecture-1 scaling factors.
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_title

        columns = self._columns()

        # Header row.
        for col_index, (header, _attr, _fmt) in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        # Data rows.
        for row_offset, data_row in enumerate(rows, start=2):
            for col_index, (_header, attribute, number_format) in enumerate(
                columns, start=1
            ):
                value = self._cell_value(data_row, attribute)
                cell = worksheet.cell(row=row_offset, column=col_index, value=value)
                if number_format is not None and value is not None and value != "":
                    cell.number_format = number_format

        self._apply_layout(worksheet, columns)

        if scaling_rows is not None:
            self._write_scaling_sheet(workbook, scaling_rows)

        workbook.save(file_path)

    # ----------------------------------------------------------------------
    def _write_scaling_sheet(self, workbook, scaling_rows: list[ScalingFactorRow]) -> None:
        """Add a compact sheet listing the model/full-scale scaling factors."""
        sheet = workbook.create_sheet(title="Scaling (model-full)")
        headers = ("Quantity", "Combination", "Scale factor (model / full-scale)")
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
        for row_offset, factor in enumerate(scaling_rows, start=2):
            sheet.cell(row=row_offset, column=1, value=factor.quantity)
            sheet.cell(row=row_offset, column=2, value=factor.combination)
            value_cell = sheet.cell(row=row_offset, column=3, value=factor.value)
            value_cell.number_format = "0.000E+00"
        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 30
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 16
        sheet.column_dimensions["C"].width = 32

    # ----------------------------------------------------------------------
    def _apply_layout(self, worksheet, columns) -> None:
        """Freeze the header, widen columns and set a sensible header height."""
        worksheet.freeze_panes = "A2"
        worksheet.row_dimensions[1].height = 42
        for col_index, (header, _attr, _fmt) in enumerate(columns, start=1):
            letter = get_column_letter(col_index)
            # Width heuristic: a bit wider than the longest word in the header.
            longest_word = max(header.split(), key=len)
            worksheet.column_dimensions[letter].width = max(10, len(longest_word) + 4)
